# =============================================================================
# Router WO Dashboard — app.py
#
# Sections
#   1. Imports
#   2. Data config         — file path, sheet, WO cell mapping
#   3. Status definitions  — SINGLE SOURCE OF TRUTH (add new statuses here)
#   4. Visual tokens & CSS
#   5. Helpers             — label lookup, HTML builders
#   6. Data layer          — load, validate, audit table
#   7. Chart builders      — stacked 100 %, donut, bar
#   8. UI                  — sidebar → header → KPIs → charts → audit
# =============================================================================

# ── 1. Imports ────────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
from datetime import datetime

# ── 2. Data config ────────────────────────────────────────────────────────────
DATA_FILE  = Path(__file__).parent / "data" / "Router Graphic Report MK 0.xlsx"
SHEET_NAME = "Contagem"
STATUS_COL = 6   # column F (1-based)
QTD_COL    = 8   # column H (1-based)
LOGO_CANDIDATES = [
    Path(__file__).parent / "logo1.jpg",
    Path(__file__).parent / "logo3.ico",
    Path(
        r"C:\Users\MatheusCosta\.cursor\projects\c-Projetos-5-Project-Test\assets\c__Projetos_5._Project_-_Test_logo1.jpg"
    ),
]

WO_CONFIG = [
    {"label": "WORK ORDER #1", "title_cell": (30, 2), "data_rows": (30, 33)},
    {"label": "WORK ORDER #2", "title_cell": (38, 2), "data_rows": (38, 41)},
    {"label": "WORK ORDER #3", "title_cell": (46, 2), "data_rows": (46, 49)},
]

# ── 3. Status definitions — single source of truth ───────────────────────────
# To add or rename a status: edit ONLY this list. Everything else auto-updates.
STATUSES = [
    {"key": "NOT STARTED", "label": "Not Started",  "color": "#3b82f6"},
    {"key": "CLOSE",       "label": "Completed",    "color": "#16a34a"},
    {"key": "WIP",         "label": "In Progress",  "color": "#d97706"},
    {"key": "CANCELED",    "label": "Canceled",     "color": "#dc2626"},
]

# Derived lookups — auto-generated from STATUSES, do not edit manually
STATUS_ORDER  = [s["key"]          for s in STATUSES]
STATUS_LABELS = {s["key"]: s["label"] for s in STATUSES}
STATUS_COLORS = {s["key"]: s["color"] for s in STATUSES}

# ── 4. Visual tokens & CSS ────────────────────────────────────────────────────
_CARD_BG     = "#ffffff"
_CARD_BORDER = "#e5e7eb"
_CARD_SHADOW = "0 1px 4px rgba(0,0,0,0.07)"
_PAGE_BG     = "#f8f9fb"
_TEXT_MAIN   = "#111827"
_TEXT_MUTED  = "#6b7280"

_CSS = f"""
<style>
:root {{
    --text-xs:   0.70rem;
    --text-sm:   0.80rem;
    --text-base: 0.95rem;
    --text-lg:   1.25rem;
    --text-xl:   2.00rem;
    --text-2xl:  1.65rem;
}}
[data-testid="stAppViewContainer"] > .main {{
    background-color: {_PAGE_BG};
}}
[data-testid="stSidebar"] > div:first-child {{
    background-color: {_CARD_BG};
    border-right: 1px solid {_CARD_BORDER};
}}
[data-testid="stSidebarContent"] {{
    padding-top: 0.08rem;
}}
[data-testid="stSidebarUserContent"] {{
    padding-top: 0rem;
    padding-bottom: 0.1rem;
}}
[data-testid="stSidebarUserContent"] hr {{
    margin: 0.4rem 0;
}}
[data-testid="stSidebarUserContent"] .stMarkdown p {{
    margin-bottom: 0.1rem;
}}
[data-testid="stSidebarUserContent"] .stCaptionContainer {{
    margin-bottom: 0.08rem;
}}
[data-testid="stSidebarUserContent"] h3 {{
    margin: 0.05rem 0 0.25rem 0;
}}
[data-testid="stSidebarUserContent"] [data-testid="stImage"] {{
    margin: 0 auto 0.12rem auto;
    text-align: center;
}}
[data-testid="stSidebarUserContent"] [data-testid="stImage"] img {{
    max-width: 90px;
    height: auto;
}}
.dash-header {{
    display: flex;
    align-items: baseline;
    justify-content: space-between;
    padding-bottom: 0.4rem;
}}
.dash-title {{
    font-size: var(--text-2xl);
    font-weight: 800;
    color: {_TEXT_MAIN};
    margin: 0;
    letter-spacing: -0.01em;
}}
.dash-timestamp {{
    font-size: var(--text-sm);
    color: {_TEXT_MUTED};
}}
.section-label {{
    font-size: var(--text-xs);
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: {_TEXT_MUTED};
    margin-bottom: 0.15rem;
}}
.global-tile {{
    background: {_CARD_BG};
    border: 1px solid {_CARD_BORDER};
    border-radius: 10px;
    padding: 0.85rem 1.1rem;
    text-align: center;
}}
.global-tile-value {{
    font-size: var(--text-xl);
    font-weight: 800;
    color: {_TEXT_MAIN};
    line-height: 1.1;
}}
.global-tile-label {{
    font-size: var(--text-xs);
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.07em;
    color: {_TEXT_MUTED};
    margin-top: 0.15rem;
}}
.global-tile-sub {{
    font-size: var(--text-sm);
    color: #374151;
    margin-top: 0.1rem;
    font-weight: 500;
}}
.kpi-card {{
    background: {_CARD_BG};
    border: 1px solid {_CARD_BORDER};
    border-radius: 12px;
    box-shadow: {_CARD_SHADOW};
    padding: 1.15rem 1.35rem 1rem;
}}
.kpi-wo-tag {{
    font-size: var(--text-xs);
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: {_TEXT_MUTED};
    margin-bottom: 0.1rem;
}}
.kpi-wo-title {{
    font-size: var(--text-sm);
    color: #374151;
    font-weight: 500;
    margin-bottom: 0.85rem;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}}
.kpi-main {{
    display: flex;
    align-items: flex-end;
    gap: 1.2rem;
    margin-bottom: 0.55rem;
}}
.kpi-pct-num {{
    font-size: var(--text-xl);
    font-weight: 800;
    color: #16a34a;
    line-height: 1;
}}
.kpi-pct-sub {{
    font-size: var(--text-xs);
    color: {_TEXT_MUTED};
    margin-top: 0.15rem;
}}
.kpi-total-num {{
    font-size: var(--text-lg);
    font-weight: 700;
    color: {_TEXT_MAIN};
    line-height: 1;
}}
.kpi-total-sub {{
    font-size: var(--text-xs);
    color: {_TEXT_MUTED};
    margin-top: 0.15rem;
}}
.progress-track {{
    display: flex;
    height: 9px;
    border-radius: 5px;
    overflow: hidden;
    margin: 0.7rem 0 0.8rem;
    background: #f3f4f6;
    gap: 1px;
}}
.kpi-breakdown {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 0.22rem 0.4rem;
}}
.kpi-row {{
    display: flex;
    align-items: center;
    gap: 0.35rem;
    font-size: var(--text-xs);
    color: #374151;
}}
.kpi-dot {{
    width: 8px;
    height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
}}
.kpi-count {{
    font-weight: 600;
    color: {_TEXT_MAIN};
}}
</style>
"""

# ── 5. Helpers ────────────────────────────────────────────────────────────────

def label(key: str) -> str:
    return STATUS_LABELS.get(key, key)

def make_kpi_card(wo_label: str, wo_title: str, stats: dict, total: int) -> str:
    completed_key = "CLOSE"
    pct = round(stats[completed_key] / total * 100, 1) if total > 0 else 0

    segs = "".join(
        f'<div style="width:{round(stats[s["key"]]/total*100,2) if total else 0}%;'
        f'background:{s["color"]};height:100%;"></div>'
        for s in STATUSES
    )
    rows = "".join(
        f'<div class="kpi-row">'
        f'<div class="kpi-dot" style="background:{s["color"]}"></div>'
        f'{s["label"]}: <span class="kpi-count">{stats[s["key"]]:,}</span>'
        f'</div>'
        for s in STATUSES
    )

    return f"""
    <div class="kpi-card">
        <div class="kpi-wo-tag">{wo_label}</div>
        <div class="kpi-wo-title">{wo_title}</div>
        <div class="kpi-main">
            <div>
                <div class="kpi-pct-num">{pct}%</div>
                <div class="kpi-pct-sub">completion rate</div>
            </div>
            <div>
                <div class="kpi-total-num">{total:,}</div>
                <div class="kpi-total-sub">total tasks</div>
            </div>
        </div>
        <div class="progress-track">{segs}</div>
        <div class="kpi-breakdown">{rows}</div>
    </div>
    """

def global_tile_html(value: str, tile_label: str, sub: str = "") -> str:
    sub_html = f'<div class="global-tile-sub">{sub}</div>' if sub else ""
    return f"""
    <div class="global-tile">
        <div class="global-tile-value">{value}</div>
        <div class="global-tile-label">{tile_label}</div>
        {sub_html}
    </div>
    """

# ── 6. Data layer ─────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_data(file_path: Path) -> tuple[pd.DataFrame, datetime]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[SHEET_NAME]

    records = []
    for wo in WO_CONFIG:
        first_row, last_row = wo["data_rows"]
        title = ws.cell(row=wo["title_cell"][0], column=wo["title_cell"][1]).value or wo["label"]
        title = str(title).strip()

        for row in range(first_row, last_row + 1):
            status_raw = ws.cell(row=row, column=STATUS_COL).value
            qtd_raw    = ws.cell(row=row, column=QTD_COL).value

            status = str(status_raw).strip().upper() if status_raw is not None else ""
            qtd    = int(qtd_raw) if isinstance(qtd_raw, (int, float)) else 0

            records.append({"wo": wo["label"], "wo_title": title, "status": status, "qtd": qtd})

    df = pd.DataFrame(records)
    df["status"] = pd.Categorical(df["status"], categories=STATUS_ORDER, ordered=True)
    return df, datetime.now()


def validate(df: pd.DataFrame) -> list[str]:
    issues = []
    for wo_label, group in df.groupby("wo"):
        found    = set(group["status"].dropna().astype(str))
        expected = set(STATUS_ORDER)
        missing  = expected - found
        extra    = found - expected
        if missing:
            issues.append(f"**{wo_label}**: missing statuses — {[label(s) for s in sorted(missing)]}")
        if extra:
            issues.append(f"**{wo_label}**: unrecognized statuses — {sorted(extra)}")
        if (group["qtd"] < 0).any():
            issues.append(f"**{wo_label}**: negative quantity found")
        if group["qtd"].isna().any():
            issues.append(f"**{wo_label}**: empty quantity cell")
    return issues


def build_audit_table(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(
        index="status", columns="wo", values="qtd", aggfunc="sum", observed=False
    )
    pivot = pivot.reindex(STATUS_ORDER)
    pivot.index = [label(s) for s in pivot.index]
    pivot.loc["**TOTAL**"] = pivot.sum()
    pivot = pivot.fillna(0).astype(int)
    pivot.index.name = "Status"
    return pivot

# ── 7. Chart builders ─────────────────────────────────────────────────────────

def make_stacked_100(df: pd.DataFrame, wo_labels: list[str]) -> go.Figure:
    fig   = go.Figure()
    short = [w.replace("WORK ORDER ", "WO ") for w in wo_labels]

    # Collect data indexed by WO so annotations can reference per-WO totals
    wo_totals: list[int] = []
    by_status: dict[str, tuple[list[float], list[int]]] = {}

    for s in STATUSES:
        by_status[s["key"]] = ([], [])

    for wo_label in wo_labels:
        wo_df = df[df["wo"] == wo_label]
        total = int(wo_df["qtd"].sum())
        wo_totals.append(total)
        for s in STATUSES:
            qty = int(wo_df.loc[wo_df["status"] == s["key"], "qtd"].sum())
            by_status[s["key"]][0].append(round(qty / total * 100, 1) if total > 0 else 0)
            by_status[s["key"]][1].append(qty)

    for s in STATUSES:
        pcts, counts = by_status[s["key"]]
        fig.add_trace(go.Bar(
            name=s["label"],
            x=short,
            y=pcts,
            marker=dict(color=s["color"], line=dict(color="white", width=1.5)),
            text=[f"{p}%" for p in pcts],
            textposition="inside",
            textfont=dict(size=12, color="white", family="sans-serif"),
            customdata=counts,
            hovertemplate=(
                f"<b>{s['label']}</b><br>"
                "%{x}: %{y:.1f}%<br>"
                "Count: %{customdata:,}<extra></extra>"
            ),
        ))

    fig.update_layout(
        barmode="stack",
        title=dict(text="Status Distribution Comparison", x=0.5, xanchor="center",
                   font=dict(size=14, color=_TEXT_MAIN)),
        xaxis=dict(title="", fixedrange=True),
        yaxis=dict(title="Share (%)", range=[0, 112], ticksuffix="%",
                   dtick=20, gridcolor="#e8e8e8", fixedrange=True),
        legend=dict(orientation="h", yanchor="bottom", y=1.04,
                    xanchor="center", x=0.5, font=dict(size=11)),
        uniformtext=dict(minsize=10, mode="hide"),
        plot_bgcolor="white",
        paper_bgcolor="rgba(0,0,0,0)",
        bargap=0.35,
        margin=dict(t=80, b=30, l=55, r=20),
        height=420,
    )

    # Total (n=X) above each bar for absolute context
    for wo_short, total in zip(short, wo_totals):
        fig.add_annotation(
            x=wo_short, y=103,
            text=f"n={total:,}",
            showarrow=False,
            font=dict(size=10, color="#6b7280"),
            yanchor="bottom",
        )

    return fig


def make_gauge(pct: float, title: str) -> go.Figure:
    if pct >= 80:
        bar_color = "#16a34a"
    elif pct >= 50:
        bar_color = "#d97706"
    else:
        bar_color = "#dc2626"

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=pct,
        number={"suffix": "%", "font": {"size": 42, "color": bar_color}, "valueformat": ".1f"},
        gauge={
            "axis": {"range": [0, 100], "tickwidth": 1, "tickcolor": "#e5e7eb",
                     "tickfont": {"size": 10, "color": "#9ca3af"}},
            "bar": {"color": bar_color, "thickness": 0.28},
            "bgcolor": "white",
            "borderwidth": 0,
            "steps": [
                {"range": [0,  50], "color": "#fee2e2"},
                {"range": [50, 80], "color": "#fef3c7"},
                {"range": [80, 100], "color": "#dcfce7"},
            ],
            "threshold": {
                "line": {"color": bar_color, "width": 3},
                "thickness": 0.8,
                "value": pct,
            },
        },
    ))
    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor="center",
                   font=dict(size=13, color=_TEXT_MAIN)),
        margin=dict(t=50, b=10, l=30, r=30),
        height=260,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig


def make_bar(wo_df: pd.DataFrame, title: str) -> go.Figure:
    labels_disp = [label(s) for s in wo_df["status"]]
    colors      = [STATUS_COLORS[s] for s in wo_df["status"]]
    total       = wo_df["qtd"].sum()
    y_max       = int(wo_df["qtd"].max())

    text_labels, text_positions, text_colors = [], [], []
    for v in wo_df["qtd"]:
        pct = v / total * 100 if total > 0 else 0
        if v >= y_max * 0.15:
            text_labels.append(f"{v:,}")
            text_positions.append("inside")
            text_colors.append("white")
        else:
            text_labels.append(f"{v:,} ({pct:.1f}%)")
            text_positions.append("outside")
            text_colors.append(_TEXT_MAIN)

    fig = go.Figure(go.Bar(
        x=labels_disp,
        y=wo_df["qtd"],
        marker=dict(color=colors, line=dict(color="white", width=1)),
        text=text_labels,
        textposition=text_positions,
        textfont=dict(size=11, color=text_colors),
        hovertemplate="<b>%{x}</b><br>Count: %{y:,}<extra></extra>",
    ))

    fig.update_layout(
        title=dict(text=title, x=0.5, xanchor="center",
                   font=dict(size=13, color=_TEXT_MAIN)),
        xaxis=dict(title="", tickangle=-30, fixedrange=True),
        yaxis=dict(title="Count", gridcolor="#e8e8e8", fixedrange=True,
                   range=[0, y_max * 1.30]),
        plot_bgcolor="white",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=50, b=80, l=50, r=20),
        height=380,
    )
    return fig

# ── 8. UI ─────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Router WO Dashboard", page_icon="✈️", layout="wide")
st.markdown(_CSS, unsafe_allow_html=True)

# ── Sidebar: static controls (defined before data load) ──
with st.sidebar:
    logo_path = next((p for p in LOGO_CANDIDATES if p.exists()), None)
    if logo_path is not None:
        st.image(str(logo_path), width=84)

    st.markdown("### Control Panel")
    st.divider()
    if st.button("Reload data", width="stretch", type="primary"):
        st.cache_data.clear()
        st.rerun()
    st.markdown('<p class="section-label">Filters</p>', unsafe_allow_html=True)
    wo_options  = ["All WOs"] + [wo["label"] for wo in WO_CONFIG]
    selected_wo = st.selectbox("Work Order", wo_options, label_visibility="collapsed")

# ── Guard: file missing ──
if not DATA_FILE.exists():
    st.error(
        f"File not found: `{DATA_FILE.name}`. "
        "Add the Excel file to the `data/` folder and reload.",
        icon="❌",
    )
    st.stop()

# ── Load data with error handling & spinner ──
try:
    with st.spinner("Loading data…"):
        df, loaded_at = load_data(DATA_FILE)
except KeyError:
    st.error(
        f"Sheet `{SHEET_NAME}` not found in `{DATA_FILE.name}`. "
        "Check that the file has not been restructured.",
        icon="❌",
    )
    st.stop()
except Exception as exc:
    st.error(f"Failed to read the Excel file: `{exc}`", icon="❌")
    st.info(
        "Possible causes: file is open in Excel, sheet is missing, or the file is corrupted. "
        "Fix the issue and click **Reload data**.",
        icon="💡",
    )
    st.stop()

# ── Sidebar: timestamp (needs loaded_at from data load) ──
with st.sidebar:
    st.divider()
    st.markdown('<p class="section-label">Source file</p>', unsafe_allow_html=True)
    source_caption = DATA_FILE.name if len(DATA_FILE.name) <= 24 else f"{DATA_FILE.name[:21]}..."
    st.caption(f"`{source_caption}`")
    st.markdown('<p class="section-label">Last loaded</p>', unsafe_allow_html=True)
    st.caption(loaded_at.strftime("%m/%d/%Y at %H:%M:%S"))

# ── Resolve filters ──
wo_labels     = [wo["label"] for wo in WO_CONFIG]
single_wo     = selected_wo != "All WOs"
active_labels = [selected_wo] if single_wo else wo_labels
active_df     = df[df["wo"].isin(active_labels)]
n_cols        = len(active_labels)

# ── Header ──
st.markdown(
    f'<div class="dash-header">'
    f'<p class="dash-title">Router WO Dashboard</p>'
    f'<span class="dash-timestamp">Last updated: {loaded_at.strftime("%m/%d/%Y at %H:%M:%S")}</span>'
    f'</div>',
    unsafe_allow_html=True,
)

# ── Validation banner (always against full dataset) ──
issues = validate(df)
if issues:
    for msg in issues:
        st.warning(msg, icon="⚠️")
else:
    st.success("Data validated — all statuses present and quantities consistent.", icon="✅")

st.divider()

# ── Global KPI tiles (all WOs only) ──
if not single_wo:
    st.markdown('<p class="section-label">Overall summary</p>', unsafe_allow_html=True)

    total_all  = int(df["qtd"].sum())
    closed_all = int(df.loc[df["status"] == "CLOSE", "qtd"].sum())
    rate_all   = round(closed_all / total_all * 100, 1) if total_all > 0 else 0

    wo_rates = {}
    for wl in wo_labels:
        wo_df = df[df["wo"] == wl]
        t = int(wo_df["qtd"].sum())
        c = int(wo_df.loc[wo_df["status"] == "CLOSE", "qtd"].sum())
        wo_rates[wl] = round(c / t * 100, 1) if t > 0 else 0

    best_wo  = max(wo_rates, key=wo_rates.get)
    worst_wo = min(wo_rates, key=wo_rates.get)

    g1, g2, g3, g4 = st.columns(4)
    g1.markdown(global_tile_html(f"{total_all:,}",        "Total Tasks",        "across all WOs"),                       unsafe_allow_html=True)
    g2.markdown(global_tile_html(f"{rate_all}%",           "Overall Completion",  f"{closed_all:,} completed"),           unsafe_allow_html=True)
    g3.markdown(global_tile_html(f"{wo_rates[best_wo]}%",  "Best WO",             best_wo.replace("WORK ORDER", "WO")),   unsafe_allow_html=True)
    g4.markdown(global_tile_html(f"{wo_rates[worst_wo]}%", "Lowest WO",           worst_wo.replace("WORK ORDER", "WO")),  unsafe_allow_html=True)

    st.divider()

# ── Per-WO KPI cards ──
st.markdown('<p class="section-label">Summary by Work Order</p>', unsafe_allow_html=True)
for col, wo_label in zip(st.columns(n_cols), active_labels):
    wo_df    = active_df[active_df["wo"] == wo_label]
    wo_title = wo_df["wo_title"].iloc[0]
    total    = int(wo_df["qtd"].sum())
    stats    = {s["key"]: int(wo_df.loc[wo_df["status"] == s["key"], "qtd"].sum()) for s in STATUSES}
    col.markdown(make_kpi_card(wo_label, wo_title, stats, total), unsafe_allow_html=True)

st.divider()

# ── Status distribution: 100% stacked (all WOs) ──
if not single_wo:
    st.markdown('<p class="section-label">Status distribution by Work Order</p>', unsafe_allow_html=True)
    st.plotly_chart(make_stacked_100(df, wo_labels), width="stretch", key="stacked_100")
    st.divider()

# ── Completion rate gauges ──
st.markdown('<p class="section-label">Completion rate</p>', unsafe_allow_html=True)
for col, wo_label in zip(st.columns(n_cols), active_labels):
    wo_df  = active_df[active_df["wo"] == wo_label]
    total  = int(wo_df["qtd"].sum())
    closed = int(wo_df.loc[wo_df["status"] == "CLOSE", "qtd"].sum())
    pct    = round(closed / total * 100, 1) if total > 0 else 0
    col.plotly_chart(make_gauge(pct, wo_label), width="stretch", key=f"gauge_{wo_label}")

st.divider()

# ── Quantity by status bars ──
st.markdown('<p class="section-label">Quantity by status</p>', unsafe_allow_html=True)
for col, wo_label in zip(st.columns(n_cols), active_labels):
    wo_df = active_df[active_df["wo"] == wo_label].sort_values("status")
    col.plotly_chart(make_bar(wo_df, wo_label), width="stretch", key=f"bar_{wo_label}")

st.divider()

# ── Audit table ──
with st.expander("Audit — verify against Excel", expanded=False):
    audit = build_audit_table(active_df)
    st.dataframe(
        audit.style.highlight_max(axis=1, color="#d4edda").format("{:,}"),
        use_container_width=True,
    )
    st.caption(
        "Source: `Contagem` sheet — H30:H33 (WO #1) · H38:H41 (WO #2) · H46:H49 (WO #3). "
        "Compare the **TOTAL** row against Excel totals for a quick sanity check."
    )
